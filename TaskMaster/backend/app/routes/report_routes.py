# Report routes with export functionality
from flask import Blueprint, jsonify, make_response, request
from datetime import datetime, timedelta
from app.models import Task, Category
import csv
import io
import json

# For Excel support (optional - requires openpyxl)
try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils.dataframe import dataframe_to_rows
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False

report_bp = Blueprint('reports', __name__)

# ----------------------------
# Helper Functions
# ----------------------------
def get_tasks_in_date_range(start_date, end_date):
    """Get tasks within a specific date range"""
    tasks = Task.query.filter(Task.deadline != None).all()
    filtered_tasks = [
        t for t in tasks 
        if start_date <= t.deadline.date() <= end_date
    ]
    return filtered_tasks

def format_task_for_export(task):
    """Format task data for export"""
    return {
        'ID': task.id,
        'Title': task.title,
        'Description': task.description or '',
        'Category': task.category.name if task.category else 'Uncategorized',
        'Priority': task.priority,
        'Deadline': task.deadline.strftime("%Y-%m-%d %H:%M:%S") if task.deadline else '',
        'Status': task.status,
        'Created At': task.created_at.strftime("%Y-%m-%d %H:%M:%S") if task.created_at else '',
        'Updated At': task.updated_at.strftime("%Y-%m-%d %H:%M:%S") if task.updated_at else ''
    }

def create_csv_response(data, filename):
    """Create CSV response for download"""
    output = io.StringIO()

    if data:
        fieldnames = data[0].keys()
        writer = csv.DictWriter(output, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(data)

    response = make_response(output.getvalue())
    response.headers['Content-Disposition'] = f'attachment; filename={filename}'
    response.headers['Content-Type'] = 'text/csv'
    return response

def create_excel_response(data, filename, sheet_name):
    """Create Excel response for download"""
    if not EXCEL_AVAILABLE:
        return jsonify({'message': 'Excel export not available. Install openpyxl package.'}), 500

    output = io.BytesIO()
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = sheet_name

    if data:
        # Headers
        headers = list(data[0].keys())
        for col, header in enumerate(headers, 1):
            cell = sheet.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")

        # Data rows
        for row, item in enumerate(data, 2):
            for col, header in enumerate(headers, 1):
                sheet.cell(row=row, column=col, value=item[header])

        # Auto-adjust column widths
        for column in sheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            sheet.column_dimensions[column_letter].width = adjusted_width

    workbook.save(output)
    output.seek(0)

    response = make_response(output.getvalue())
    response.headers['Content-Disposition'] = f'attachment; filename={filename}'
    response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    return response

# ----------------------------
# Weekly report (JSON)
# ----------------------------
@report_bp.route('/weekly', methods=['GET'])
def weekly_report():
    """Get weekly report in JSON format"""
    today = datetime.utcnow().date()
    week_end = today + timedelta(days=7)

    tasks = get_tasks_in_date_range(today, week_end)
    weekly_tasks = [
        {
            'id': t.id,
            'title': t.title,
            'description': t.description,
            'category': t.category.name if t.category else None,
            'priority': t.priority,
            'deadline': t.deadline.strftime("%Y-%m-%d %H:%M:%S"),
            'status': t.status
        }
        for t in tasks
    ]

    # Calculate statistics
    completed_count = len([t for t in tasks if t.status == 'Completed'])
    pending_count = len([t for t in tasks if t.status == 'Pending'])

    return jsonify({
        'week_start': today.strftime("%Y-%m-%d"),
        'week_end': week_end.strftime("%Y-%m-%d"),
        'tasks': weekly_tasks,
        'total_tasks': len(weekly_tasks),
        'completed_tasks': completed_count,
        'pending_tasks': pending_count,
        'completion_rate': round((completed_count / len(tasks) * 100) if len(tasks) > 0 else 0, 1)
    }), 200

# ----------------------------
# Weekly report export (CSV)
# ----------------------------
@report_bp.route('/weekly/export/csv', methods=['GET'])
def weekly_report_csv():
    """Export weekly report as CSV"""
    today = datetime.utcnow().date()
    week_end = today + timedelta(days=7)

    tasks = get_tasks_in_date_range(today, week_end)
    export_data = [format_task_for_export(task) for task in tasks]

    filename = f'weekly_report_{today.strftime("%Y%m%d")}.csv'
    return create_csv_response(export_data, filename)

# ----------------------------
# Weekly report export (Excel)
# ----------------------------
@report_bp.route('/weekly/export/excel', methods=['GET'])
def weekly_report_excel():
    """Export weekly report as Excel"""
    today = datetime.utcnow().date()
    week_end = today + timedelta(days=7)

    tasks = get_tasks_in_date_range(today, week_end)
    export_data = [format_task_for_export(task) for task in tasks]

    filename = f'weekly_report_{today.strftime("%Y%m%d")}.xlsx'
    sheet_name = f'Week {today.strftime("%Y-%m-%d")}'
    return create_excel_response(export_data, filename, sheet_name)

# ----------------------------
# Monthly report (JSON)
# ----------------------------
@report_bp.route('/monthly', methods=['GET'])
def monthly_report():
    """Get monthly report in JSON format"""
    today = datetime.utcnow().date()
    month_start = today.replace(day=1)
    month_end = (month_start + timedelta(days=32)).replace(day=1) - timedelta(days=1)

    tasks = get_tasks_in_date_range(month_start, month_end)
    monthly_tasks = [
        {
            'id': t.id,
            'title': t.title,
            'description': t.description,
            'category': t.category.name if t.category else None,
            'priority': t.priority,
            'deadline': t.deadline.strftime("%Y-%m-%d %H:%M:%S"),
            'status': t.status
        }
        for t in tasks
    ]

    # Calculate statistics
    completed_count = len([t for t in tasks if t.status == 'Completed'])
    pending_count = len([t for t in tasks if t.status == 'Pending'])

    # Category breakdown
    category_stats = {}
    for task in tasks:
        cat_name = task.category.name if task.category else 'Uncategorized'
        if cat_name not in category_stats:
            category_stats[cat_name] = {'total': 0, 'completed': 0, 'pending': 0}
        category_stats[cat_name]['total'] += 1
        if task.status == 'Completed':
            category_stats[cat_name]['completed'] += 1
        else:
            category_stats[cat_name]['pending'] += 1

    return jsonify({
        'month_start': month_start.strftime("%Y-%m-%d"),
        'month_end': month_end.strftime("%Y-%m-%d"),
        'tasks': monthly_tasks,
        'total_tasks': len(monthly_tasks),
        'completed_tasks': completed_count,
        'pending_tasks': pending_count,
        'completion_rate': round((completed_count / len(tasks) * 100) if len(tasks) > 0 else 0, 1),
        'category_breakdown': category_stats
    }), 200

# ----------------------------
# Monthly report export (CSV)
# ----------------------------
@report_bp.route('/monthly/export/csv', methods=['GET'])
def monthly_report_csv():
    """Export monthly report as CSV"""
    today = datetime.utcnow().date()
    month_start = today.replace(day=1)
    month_end = (month_start + timedelta(days=32)).replace(day=1) - timedelta(days=1)

    tasks = get_tasks_in_date_range(month_start, month_end)
    export_data = [format_task_for_export(task) for task in tasks]

    filename = f'monthly_report_{today.strftime("%Y%m")}.csv'
    return create_csv_response(export_data, filename)

# ----------------------------
# Monthly report export (Excel)
# ----------------------------
@report_bp.route('/monthly/export/excel', methods=['GET'])
def monthly_report_excel():
    """Export monthly report as Excel"""
    today = datetime.utcnow().date()
    month_start = today.replace(day=1)
    month_end = (month_start + timedelta(days=32)).replace(day=1) - timedelta(days=1)

    tasks = get_tasks_in_date_range(month_start, month_end)
    export_data = [format_task_for_export(task) for task in tasks]

    filename = f'monthly_report_{today.strftime("%Y%m")}.xlsx'
    sheet_name = f'Month {today.strftime("%Y-%m")}'
    return create_excel_response(export_data, filename, sheet_name)

# ----------------------------
# Custom date range report
# ----------------------------
@report_bp.route('/custom', methods=['GET'])
def custom_report():
    """Get report for custom date range"""
    start_date_str = request.args.get('start_date')
    end_date_str = request.args.get('end_date')

    if not start_date_str or not end_date_str:
        return jsonify({'message': 'Both start_date and end_date are required (YYYY-MM-DD)'}), 400

    try:
        start_date = datetime.strptime(start_date_str, "%Y-%m-%d").date()
        end_date = datetime.strptime(end_date_str, "%Y-%m-%d").date()
    except ValueError:
        return jsonify({'message': 'Invalid date format. Use YYYY-MM-DD'}), 400

    if start_date > end_date:
        return jsonify({'message': 'Start date must be before or equal to end date'}), 400

    tasks = get_tasks_in_date_range(start_date, end_date)
    custom_tasks = [
        {
            'id': t.id,
            'title': t.title,
            'description': t.description,
            'category': t.category.name if t.category else None,
            'priority': t.priority,
            'deadline': t.deadline.strftime("%Y-%m-%d %H:%M:%S"),
            'status': t.status
        }
        for t in tasks
    ]

    # Calculate statistics
    completed_count = len([t for t in tasks if t.status == 'Completed'])
    pending_count = len([t for t in tasks if t.status == 'Pending'])

    return jsonify({
        'date_range_start': start_date.strftime("%Y-%m-%d"),
        'date_range_end': end_date.strftime("%Y-%m-%d"),
        'tasks': custom_tasks,
        'total_tasks': len(custom_tasks),
        'completed_tasks': completed_count,
        'pending_tasks': pending_count,
        'completion_rate': round((completed_count / len(tasks) * 100) if len(tasks) > 0 else 0, 1)
    }), 200

# ----------------------------
# Custom date range export (CSV)
# ----------------------------
@report_bp.route('/custom/export/csv', methods=['GET'])
def custom_report_csv():
    """Export custom date range report as CSV"""
    start_date_str = request.args.get('start_date')
    end_date_str = request.args.get('end_date')

    if not start_date_str or not end_date_str:
        return jsonify({'message': 'Both start_date and end_date are required (YYYY-MM-DD)'}), 400

    try:
        start_date = datetime.strptime(start_date_str, "%Y-%m-%d").date()
        end_date = datetime.strptime(end_date_str, "%Y-%m-%d").date()
    except ValueError:
        return jsonify({'message': 'Invalid date format. Use YYYY-MM-DD'}), 400

    if start_date > end_date:
        return jsonify({'message': 'Start date must be before or equal to end date'}), 400

    tasks = get_tasks_in_date_range(start_date, end_date)
    export_data = [format_task_for_export(task) for task in tasks]

    filename = f'custom_report_{start_date.strftime("%Y%m%d")}_{end_date.strftime("%Y%m%d")}.csv'
    return create_csv_response(export_data, filename)

# ----------------------------
# Custom date range export (Excel)
# ----------------------------
@report_bp.route('/custom/export/excel', methods=['GET'])
def custom_report_excel():
    """Export custom date range report as Excel"""
    start_date_str = request.args.get('start_date')
    end_date_str = request.args.get('end_date')

    if not start_date_str or not end_date_str:
        return jsonify({'message': 'Both start_date and end_date are required (YYYY-MM-DD)'}), 400

    try:
        start_date = datetime.strptime(start_date_str, "%Y-%m-%d").date()
        end_date = datetime.strptime(end_date_str, "%Y-%m-%d").date()
    except ValueError:
        return jsonify({'message': 'Invalid date format. Use YYYY-MM-DD'}), 400

    if start_date > end_date:
        return jsonify({'message': 'Start date must be before or equal to end date'}), 400

    tasks = get_tasks_in_date_range(start_date, end_date)
    export_data = [format_task_for_export(task) for task in tasks]

    filename = f'custom_report_{start_date.strftime("%Y%m%d")}_{end_date.strftime("%Y%m%d")}.xlsx'
    sheet_name = f'Report {start_date.strftime("%Y-%m-%d")} to {end_date.strftime("%Y-%m-%d")}'
    return create_excel_response(export_data, filename, sheet_name)

# ----------------------------
# All tasks export
# ----------------------------
@report_bp.route('/all/export/csv', methods=['GET'])
def all_tasks_csv():
    """Export all tasks as CSV"""
    tasks = Task.query.all()
    export_data = [format_task_for_export(task) for task in tasks]

    filename = f'all_tasks_{datetime.utcnow().strftime("%Y%m%d")}.csv'
    return create_csv_response(export_data, filename)

@report_bp.route('/all/export/excel', methods=['GET'])
def all_tasks_excel():
    """Export all tasks as Excel"""
    tasks = Task.query.all()
    export_data = [format_task_for_export(task) for task in tasks]

    filename = f'all_tasks_{datetime.utcnow().strftime("%Y%m%d")}.xlsx'
    sheet_name = 'All Tasks'
    return create_excel_response(export_data, filename, sheet_name)
